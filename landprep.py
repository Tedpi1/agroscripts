import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def fetch_data():
    # --- Connect to DB ---
    conn = mysql.connector.connect(
        host="localhost",   # Change if needed
        user="root",        # Change to your MySQL user
        password="",        # Change to your MySQL password
        database="test_db"
    )
    cursor = conn.cursor(dictionary=True)

    # --- Run Query ---
    query = """
    SELECT event_id,
           CONCAT(block_name, " ", partion_name) AS block_name,
           crop, color_code, action_type,
           WEEK(start_date, 1) AS weeks,
           YEAR(start_date) AS years,
           DATE_FORMAT(start_date, "%d %b %Y") AS start_date,
           start_date AS exert_date,
           ready_for_planting
    FROM event_lines
    LEFT JOIN events USING(event_id)
    LEFT JOIN company_map ON company_map.id = events.block_id
    LEFT JOIN blocks ON blocks.block_id = company_map.block_id
    LEFT JOIN crops USING(crop_id)
    LEFT JOIN crops_yield_project USING(crop_id)
    LEFT JOIN land_prep_batches USING(event_id)
    WHERE start_date >= DATE_SUB(CURRENT_DATE(), INTERVAL 7 DAY)
      AND action_type = "PLANTING"
      AND event_lines.is_active = 1
    GROUP BY CONCAT(block_name, " ", partion_name),
             crop, color_code, action_type,
             WEEK(start_date, 1), YEAR(start_date),
             start_date, event_id, exert_date, ready_for_planting
    ORDER BY exert_date ASC;
    """
    cursor.execute(query)
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    return results

def export_to_excel(data, filename="land_preparation.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Land Preparation"

    # Headers (match your table)
    headers = ["Partition", "Crop", "Planting Date", "Planting Wk-Year", "Status"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add rows
    for row in data:
        partition = row["block_name"]
        crop = row["crop"]
        planting_date = row["start_date"]
        wk_year = f"{row['weeks']}-{row['years']}"
        status = "Cleared" if row["ready_for_planting"] == 1 else "Not Cleared"

        ws.append([partition, crop, planting_date, wk_year, status])

    # Apply styling for Status column
    for row_idx in range(2, len(data) + 2):
        status_cell = ws.cell(row=row_idx, column=5)
        if status_cell.value == "Not Cleared":
            status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)
        else:
            status_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            status_cell.font = Font(color="000000", bold=True)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)
    print(f"Excel file generated: {filename}")

if __name__ == "__main__":
    data = fetch_data()
    export_to_excel(data)
