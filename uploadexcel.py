import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import calendar
import datetime

# --- DB connection ---
conn = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="test_db"   # change to your DB
)
cursor = conn.cursor(dictionary=True)

# --- Fetch events ---
query = """
    SELECT CONCAT(blocks.block_name, " ", company_map.partion_name) AS block_name,
        company_map.area AS area,
        crop, color_code, action_type,
        SUM(total_yeild) AS total_yeild,
        WEEK(start_date, 1) AS weeks,
        YEAR(start_date) AS years,
        MONTHNAME(start_date) AS month_name
    FROM event_lines
    LEFT JOIN events USING(event_id)
    LEFT JOIN company_map ON company_map.id = events.block_id
    LEFT JOIN blocks ON blocks.block_id = company_map.block_id
    LEFT JOIN crops USING(crop_id)
    WHERE start_date >= DATE_FORMAT(CONCAT(YEAR(CURDATE()), "-01-01"), "%Y-%m-%d")
    AND action_type NOT IN ("NURSERY")
    AND event_lines.is_active = 1
    GROUP BY block_name, area, crop, color_code, action_type, weeks, years, month_name
    ORDER BY block_name, years, weeks ASC;
"""
cursor.execute(query)
events = cursor.fetchall()

# --- Build structure: {block: {year: {week: {...}}}} ---
blocks = {}
years_set = set()

for row in events:
    block = row["block_name"]
    area = row["area"]
    year = row["years"]
    week = row["weeks"]

    years_set.add(year)

    if block not in blocks:
        blocks[block] = {"area": area, "data": {}}
    if year not in blocks[block]["data"]:
        blocks[block]["data"][year] = {}

    blocks[block]["data"][year][week] = row

# --- Create workbook ---
wb = Workbook()
ws = wb.active
ws.title = "Farm Activities"

# Default empty fill (for weeks with no activity)
empty_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

# --- Build headers ---
col = 1
ws.cell(1, col, "BLOCKS")
ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
ws.cell(1, col+1, "AREA")
ws.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2)
col = 3

for year in sorted(years_set):
    year_start_col = col
    
    # Total weeks (52 or 53 depending on year)
    total_weeks = datetime.date(year, 12, 28).isocalendar()[1]
    
    for month in calendar.month_name[1:]:  # Jan → Dec
        month_start_col = col
        for week in range(1, total_weeks + 1):
            monday = datetime.date.fromisocalendar(year, week, 1)
            if monday.strftime("%B") == month:
                ws.cell(3, col, f"Wk {week}")
                col += 1
        if col > month_start_col:  # only merge if this month had weeks
            ws.merge_cells(start_row=2, start_column=month_start_col, end_row=2, end_column=col-1)
            ws.cell(2, month_start_col, month)
    
    ws.merge_cells(start_row=1, start_column=year_start_col, end_row=1, end_column=col-1)
    ws.cell(1, year_start_col, str(year))

# --- Freeze panes (lock first 2 cols + top 3 rows) ---
ws.freeze_panes = "C4"

# --- Fill block rows ---
row = 4
for block, block_data in blocks.items():
    ws.cell(row, 1, block)
    ws.cell(row, 2, block_data["area"])

    col = 3
    for year in sorted(years_set):
        total_weeks = datetime.date(year, 12, 28).isocalendar()[1]
        for week in range(1, total_weeks + 1):
            val = ""
            fill = empty_fill  # default fill (gray for empty weeks)
            
            if year in block_data["data"] and week in block_data["data"][year]:
                act = block_data["data"][year][week]["action_type"]
                crop = block_data["data"][year][week]["crop"]
                yield_val = block_data["data"][year][week]["total_yeild"]
                color_code = block_data["data"][year][week]["color_code"]

                # Use DB color if available
                if color_code:
                    hex_color = color_code.replace("#", "")
                    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                else:
                    fill = empty_fill

                if act == "PLANTING":
                    val = f"P-{crop[0]}"
                elif act == "UPROOTING":
                    val = f"U-{crop[0]}"
                elif act == "HARVESTING":
                    val = yield_val

            cell = ws.cell(row, col, val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if val and (act != "HARVESTING"):
                cell.font = Font(color="FFFFFF", bold=True)

            col += 1
    row += 1

# --- Totals row ---
ws.cell(row, 1, "Total")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

col = 3
for year in sorted(years_set):
    total_weeks = datetime.date(year, 12, 28).isocalendar()[1]
    for week in range(1, total_weeks + 1):
        formula = f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{row-1})"
        ws.cell(row, col, formula)
        col += 1

# --- Legend / Key from DB (Crop + Color only, unique) ---
legend_start_row = row + 2
ws.merge_cells(start_row=legend_start_row, start_column=1, end_row=legend_start_row, end_column=2)
ws.cell(legend_start_row, 1, "LEGEND").font = Font(bold=True)

# Fetch crops + colors
cursor.execute("""
    SELECT crop, color_code 
    FROM event_lines 
    LEFT JOIN events USING(event_id)
    LEFT JOIN crops USING(crop_id)
    WHERE event_lines.is_active = 1
    AND crop IS NOT NULL
""")
legend_items = cursor.fetchall()

# Remove duplicates (crop + color)
unique_legend = {}
for item in legend_items:
    crop = item["crop"] if item["crop"] else "N/A"
    color_code = item["color_code"]
    unique_legend[crop] = color_code   # keep one per crop

# Write legend
for i, (crop, color_code) in enumerate(unique_legend.items(), start=0):
    r = legend_start_row + i + 1
    ws.cell(r, 1, crop)

    if color_code:
        hex_color = color_code.replace("#", "")
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    else:
        fill = empty_fill

    cell = ws.cell(r, 2, "")
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# --- Auto-fit column widths ---
for col_idx, col_cells in enumerate(ws.columns, 1):
    max_length = 0
    column = get_column_letter(col_idx)
    for cell in col_cells:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save file
wb.save("farm_calendar.xlsx")
print("Excel file created: farm_calendar.xlsx")
